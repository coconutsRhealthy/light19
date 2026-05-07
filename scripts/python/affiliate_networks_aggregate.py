#!/usr/bin/env python3
"""
Generates affiliate-networks-per-company.xlsx in the project root.

For each company, lists which affiliate networks have it available
(based on column A of every network sheet in src/app/services/Sales.xlsx,
plus the inferred network of any live URL in affiliate-link.service.ts),
whether it is currently live on diski.nl, and — for not-live companies —
the candidate URLs available across networks.

Run from anywhere:
    python3 scripts/python/affiliate_networks_aggregate.py
"""

import os, re
from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

SCRIPT_DIR   = os.path.dirname(os.path.abspath(__file__))
PROJECT_ROOT = os.path.dirname(os.path.dirname(SCRIPT_DIR))
SALES_PATH   = os.path.join(PROJECT_ROOT, "src", "app", "services", "Sales.xlsx")
SERVICE_PATH = os.path.join(PROJECT_ROOT, "src", "app", "services", "affiliate-link.service.ts")
OUT_PATH     = os.path.join(PROJECT_ROOT, "affiliate-networks-per-company.xlsx")

# Sheets in Sales.xlsx that aren't an affiliate network
NON_NETWORK_SHEETS = {"New partnerships", "Status"}

# Rows in column A that aren't a real brand
JUNK_PATTERNS = [r'^more on platform$', r'^nieuwe partners', r'^tradetracker\.com$']

# Per network sheet, the 0-based column index of the "Affiliate link" column
# that lines up with column A's partnership names. Sheets not listed here
# either have no URL column or no col-A partners.
URL_COL = {
    "Shopmyinfluence": 1,
    "Linkpizza":       1,
    "Awin":            2,
    "Daisycon":        1,
    "Tradetracker":    1,
    "Tradedoubler":    1,
}

# Live-URL host patterns -> affiliate network name
URL_PATTERNS = [
    (r'tidd\.ly|awin\.com|awin1\.com', 'Awin'),
    (r'pzz\.to|linkpizza', 'Linkpizza'),
    (r'smyi\.me|shopmy\.', 'Shopmyinfluence'),
    (r'jf79\.net|daisycon|glp8\.net|bdt9\.net|fr135\.net|lt45\.net|rkn3\.net|jdt8\.net|ds1\.nl', 'Daisycon'),
    (r'tradetracker|tcsl\.tt', 'Tradetracker'),
    (r'tradedoubler|tdoubler', 'Tradedoubler'),
    (r'linksynergy|rakuten', 'Rakuten'),
    (r'partnerize|prf\.hn', 'Partnerize'),
    (r'admitad|epn\.bz', 'Admitad'),
    (r'adtr\.co|adtraction', 'Adtraction'),
    (r'flexlinkspro|flexoffers', 'Flexoffers'),
    (r'webgains|wgmm\.de', 'Webgains'),
    (r'impact\.com|pxf\.io|sjv\.io|7eer\.net|evyy\.net|ojrq\.net', 'Impact'),
    (r'optimisemedia', 'Optimise Media'),
    (r'kqzyfj\.com|cj\.dotomi\.com|qksrv\.net|apmebf\.com|dpbolvw\.net|tkqlhce\.com|anrdoezrs\.net|jdoqocy\.com|ftjcfx\.com|lduhtrp\.net', 'CJ'),
    (r'stylink\.it|stylink', 'Stylink'),
    (r'moonrover', 'Moonrover'),
]


def is_junk(name):
    return any(re.match(p, name.lower().strip()) for p in JUNK_PATTERNS)


def normalize(s):
    """Group by brand identity: drop region tokens, parens, FamilyBlend tag,
    non-alphanum, and any trailing nl/be/eu still attached."""
    s = s.lower()
    s = re.sub(r'\s*\(.*?\)\s*', ' ', s)
    s = re.sub(r'\bfamily\s*blend\b', '', s, flags=re.I)
    while True:
        new = re.sub(r'[\s\-_/+,]+(nl|be|eu|benelux|bnl|de|uk|com|nederland|global|netherlands|row)\s*$', '', s)
        if new == s: break
        s = new
    s = re.sub(r'[^a-z0-9]', '', s)
    while s.endswith(('nl', 'be', 'eu')) and len(s) > 4:
        s = s[:-2]
    return s


def clean_display(s):
    """Human-readable variant: strip parens, FamilyBlend tag, trailing region tokens."""
    s = re.sub(r'\s*\([^)]*\)\s*', ' ', s).strip()
    s = re.sub(r'\bfamily\s*blend\b', '', s, flags=re.I).strip(' -_/+,')
    while True:
        new = re.sub(r'[\s\-_/+,]+(nl|be|eu|benelux|bnl|nederland|global|netherlands|row)\s*$', '', s, flags=re.I)
        if new == s: break
        s = new
    return s.strip(' -_/+,')


def network_from_url(url):
    if not url: return ''
    u = url.lower()
    for pat, name in URL_PATTERNS:
        if re.search(pat, u): return name
    return 'Direct/Other'


def load_sheet_entries():
    """Returns {norm: {networks: set, raws: [str], urls: [(network, url)]}}."""
    wb = load_workbook(SALES_PATH, read_only=True, data_only=True)
    out = {}
    for sheet in wb.sheetnames:
        if sheet in NON_NETWORK_SHEETS: continue
        ws = wb[sheet]
        url_col = URL_COL.get(sheet)
        seen = set()
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i == 0 or not row: continue
            a = row[0] if len(row) >= 1 else None
            if a is None or not str(a).strip(): continue
            name = str(a).strip()
            if is_junk(name) or name in seen: continue
            seen.add(name)
            n = normalize(name)
            if not n: continue
            e = out.setdefault(n, {'networks': set(), 'raws': [], 'urls': []})
            e['networks'].add(sheet)
            e['raws'].append(name)
            if url_col is not None and len(row) > url_col:
                v = row[url_col]
                if v and str(v).strip().startswith('http'):
                    e['urls'].append((sheet, str(v).strip()))
    return out


def load_live_links():
    """Returns {norm: (key, url, network)} from affiliate-link.service.ts."""
    with open(SERVICE_PATH) as f:
        text = f.read()
    m = re.search(r'private affiliateLinks[^{]*\{(.*?)\n  \};', text, re.DOTALL)
    body = m.group(1) if m else ''
    out = {}
    for match in re.finditer(r"^\s*'([^']+)':\s*\{\s*\n\s*url:\s*'([^']+)'", body, re.MULTILINE):
        key, url = match.group(1), match.group(2)
        n = normalize(key)
        if n: out[n] = (key, url, network_from_url(url))
    return out


def pick_display(norm, sheet_entries, live_by_norm):
    candidates = list(sheet_entries.get(norm, {}).get('raws', []))
    if norm in live_by_norm:
        candidates.append(live_by_norm[norm][0])
    cleaned = [(clean_display(c), c) for c in candidates if clean_display(c)]
    if not cleaned: return candidates[0] if candidates else norm
    cleaned.sort(key=lambda x: (len(x[0]), -sum(1 for c in x[0] if c.isupper())))
    return cleaned[0][0]


def build_rows(sheet_entries, live_by_norm):
    rows = []
    for n in set(sheet_entries) | set(live_by_norm):
        display = pick_display(n, sheet_entries, live_by_norm)
        sd = sheet_entries.get(n, {'networks': set(), 'urls': []})
        networks = set(sd['networks'])
        live = live_by_norm.get(n)
        if live:
            live_key, live_url, live_net = live
            if live_net: networks.add(live_net)
            candidates = ''
        else:
            live_key = live_url = live_net = ''
            seen = set(); pairs = []
            for net, u in sd['urls']:
                if u in seen: continue
                seen.add(u); pairs.append((net, u))
            candidates = "\n".join(f"{net}: {u}" for net, u in pairs) if pairs else "no url available"
        rows.append({
            'display':   display,
            'networks':  sorted(networks),
            'live':      bool(live),
            'live_net':  live_net,
            'live_url':  live_url,
            'candidates': candidates,
        })
    rows.sort(key=lambda r: r['display'].lower())
    return rows


# ── Styles ─────────────────────────────────────────────────────────────────────
GRAY     = PatternFill("solid", fgColor="2C3E50")
HDR      = PatternFill("solid", fgColor="34495E")
LIVE_ROW = PatternFill("solid", fgColor="D5F5E3")
DEAD_ROW = PatternFill("solid", fgColor="FDEBD0")
WHITE    = Font(name="Arial", bold=True, color="FFFFFF", size=10)
BODY     = Font(name="Arial", size=10)
ITAL     = Font(name="Arial", italic=True, size=10, color="555555")
thin     = Side(style="thin", color="CCCCCC")
BORDER   = Border(left=thin, right=thin, top=thin, bottom=thin)


def style(cell, fill=None, font=None, align="left", wrap=False):
    if fill: cell.fill = fill
    cell.font = font or BODY
    cell.border = BORDER
    cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=wrap)


def build_xlsx(rows):
    wb = Workbook()
    ws = wb.active
    ws.title = "Per-company affiliate map"
    ws.sheet_view.showGridLines = False

    widths = [6, 36, 42, 8, 12, 14, 55, 70]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Title
    ws.merge_cells("A1:H1")
    t = ws["A1"]; t.value = "Affiliate networks per company"
    t.fill = GRAY; t.font = Font(name="Arial", bold=True, color="FFFFFF", size=13)
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    total = len(rows)
    live_count = sum(1 for r in rows if r['live'])
    ws.merge_cells("A2:H2")
    ws["A2"] = f"{total} companies — {live_count} live on diski.nl, {total - live_count} not yet live"
    ws["A2"].font = ITAL
    ws["A2"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 18

    HDR_ROW = 4
    headers = ["#", "Company", "Available at Networks", "# Networks",
               "Live on site", "Live network", "Live URL", "Candidate URLs (if not live)"]
    for col, h in enumerate(headers, start=1):
        c = ws.cell(row=HDR_ROW, column=col, value=h)
        style(c, fill=HDR, font=WHITE, align="center")
    ws.row_dimensions[HDR_ROW].height = 22
    ws.freeze_panes = "B5"

    for i, r in enumerate(rows, start=1):
        rr = HDR_ROW + i
        fill = LIVE_ROW if r['live'] else DEAD_ROW
        line_count = (r['candidates'].count('\n') + 1) if (r['candidates'] and not r['live']) else 1
        cells = [
            (i, "center", False),
            (r['display'], "left", False),
            (", ".join(r['networks']) if r['networks'] else "—", "left", False),
            (len(r['networks']), "center", False),
            ("Yes" if r['live'] else "No", "center", False),
            (r['live_net'], "left", False),
            (r['live_url'], "left", False),
            (r['candidates'], "left", True),
        ]
        for col, (val, align, wrap) in enumerate(cells, start=1):
            c = ws.cell(row=rr, column=col, value=val)
            style(c, fill=fill, align=align, wrap=wrap)
        ws.row_dimensions[rr].height = max(16, 15 * line_count)

    wb.save(OUT_PATH)
    return total, live_count


if __name__ == "__main__":
    sheet_entries = load_sheet_entries()
    live_by_norm  = load_live_links()
    rows          = build_rows(sheet_entries, live_by_norm)
    total, live   = build_xlsx(rows)
    print(f"Saved: {OUT_PATH}")
    print(f"  Total companies: {total}")
    print(f"  Live on site:    {live}  ({live/total*100:.1f}%)")
    print(f"  Not yet live:    {total - live}")
