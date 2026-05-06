#!/usr/bin/env python3
"""
Generates seo-discount-comparison.xlsx in the project root.
Lists companies present on diski.nl (discounts.json) that do NOT yet have
an SEO content .ts file in src/app/company-codes/company-seo-content/.

Companies whose SEO file exists under a slightly different name (TLD or
punctuation differences) are treated as already covered and excluded.

Run from anywhere:
    python3 scripts/python/seo_discount_comparison.py
"""

import json, os, re
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

SCRIPT_DIR   = os.path.dirname(os.path.abspath(__file__))
PROJECT_ROOT = os.path.dirname(os.path.dirname(SCRIPT_DIR))
DISCOUNTS_PATH = os.path.join(PROJECT_ROOT, "src", "app", "data", "discounts.json")
SEO_DIR        = os.path.join(PROJECT_ROOT, "src", "app", "company-codes", "company-seo-content")
OUT_PATH       = os.path.join(PROJECT_ROOT, "seo-discount-comparison.xlsx")


def normalize(name):
    n = name.lower()
    n = re.sub(r'\.(nl|com|eu|de|be|co|uk|fr|se|dk|no)$', '', n)
    n = re.sub(r'[^a-z0-9]', '', n)
    return n


def load_data():
    with open(DISCOUNTS_PATH) as f:
        entries = json.load(f)

    discount_companies = set()
    for entry in entries:
        raw  = entry.split(',')[0].strip()
        name = re.sub(r'\s*\(.*?\)', '', raw).strip()
        discount_companies.add(name)

    seo_companies = set()
    for fname in os.listdir(SEO_DIR):
        if fname.endswith('.ts') and fname != 'index.ts':
            seo_companies.add(fname[:-3])

    return discount_companies, seo_companies


def find_missing_seo(discount_companies, seo_companies):
    seo_norm = {normalize(s) for s in seo_companies}
    missing = [d for d in discount_companies if d not in seo_companies and normalize(d) not in seo_norm]
    return sorted(missing, key=str.lower)


# ── Styles ─────────────────────────────────────────────────────────────────────
RED_HDR   = PatternFill("solid", fgColor="E74C3C")
RED_ROW   = PatternFill("solid", fgColor="FADBD8")
GRAY_HDR  = PatternFill("solid", fgColor="2C3E50")
WHITE_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=10)
BODY_FONT  = Font(name="Arial", size=10)
thin   = Side(style="thin", color="CCCCCC")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)


def style_cell(cell, fill=None, font=None, align="left"):
    if fill: cell.fill = fill
    cell.font = font or BODY_FONT
    cell.border = BORDER
    cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=False)


def build_xlsx(missing_seo, total_discount, total_seo):
    wb = Workbook()
    ws = wb.active
    ws.title = "Missing SEO"
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["B"].width = 40

    ws.merge_cells("A1:B1")
    t = ws["A1"]
    t.value = "Companies on diski.nl without SEO content"
    t.fill = GRAY_HDR
    t.font = Font(name="Arial", bold=True, color="FFFFFF", size=13)
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    ws["A2"] = f"Missing: {len(missing_seo)} of {total_discount} on diski.nl ({total_seo} SEO files exist)"
    ws.merge_cells("A2:B2")
    ws["A2"].font = Font(name="Arial", italic=True, size=10, color="555555")
    ws["A2"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 18

    for col, h in enumerate(["#", "Company"], start=1):
        c = ws.cell(row=4, column=col, value=h)
        style_cell(c, fill=RED_HDR, font=WHITE_FONT, align="center")
    ws.row_dimensions[4].height = 20
    ws.freeze_panes = "A5"

    for i, name in enumerate(missing_seo, start=1):
        r = 4 + i
        c1 = ws.cell(row=r, column=1, value=i)
        c2 = ws.cell(row=r, column=2, value=name)
        style_cell(c1, fill=RED_ROW, align="center")
        style_cell(c2, fill=RED_ROW)
        ws.row_dimensions[r].height = 16

    wb.save(OUT_PATH)
    print(f"Saved: {OUT_PATH}")
    print(f"  Missing SEO: {len(missing_seo)} of {total_discount} companies on diski.nl")


if __name__ == "__main__":
    discount_companies, seo_companies = load_data()
    missing_seo = find_missing_seo(discount_companies, seo_companies)
    build_xlsx(missing_seo, len(discount_companies), len(seo_companies))
